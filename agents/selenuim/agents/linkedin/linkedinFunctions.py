import re
import sys
import keyboard
import openpyxl
import os

from htmlrag import clean_html
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium import webdriver
import time
from datetime import datetime
from selenium.common.exceptions import NoSuchElementException, TimeoutException, NoSuchDriverException, \
    WebDriverException, ElementNotVisibleException, ElementNotInteractableException


def createNameAndHyperLinkLists(xlsxFilePath):
    """
    Extract data from specified columns in an Excel file and create lists for further processing.

    This function loads the Excel workbook, iterates through the rows in the active sheet, and extracts
    information such as first names, last names, hyperlinks, statuses, and time approached. The extracted
    data is then stored in separate lists.

    Parameters:
    - xlsxFilePath (str): Path to the Excel file.

    Returns:
    tuple: A tuple containing lists of first names, last names, hyperlinks, statuses, and time approached.
    """
    try:
        wb = openpyxl.load_workbook(xlsxFilePath)
        ws = wb.active

        # Create the lists to store the values
        firstNames = []
        lastNames = []
        nameLinks = []
        statuses = []
        timeApproachedList = []

        # Iterate over the rows in the sheet
        for i, row in enumerate(ws):
            if row[0].value is not None:
                name = row[0].value
                # print(name)
                firsName, lastName = splitFirstNameAndLastName(name)
                Link = ws.cell(row=i + 1, column=1).hyperlink.target
                status = ws.cell(row=i + 1, column=2).value
                timeApproached = ws.cell(row=i + 1, column=3).value
                # Add the values to the lists
                firstNames.append(firsName)
                lastNames.append(lastName)
                nameLinks.append(Link)
                statuses.append(status)
                timeApproachedList.append(timeApproached)
            else:
                wb.close()
                break
        wb.close()
        return firstNames, lastNames, nameLinks, statuses, timeApproachedList
    except Exception as e:
        print("Exception In function 'createNameAndHyperLinkLists'")
        print(e)


def splitFirstNameAndLastName(name: str):
    """
    Split a full name into first and last names.

    This function takes a full name as input and splits it into two parts: the first name and the last name.
    If the full name contains only one part, the last name will be an empty string.

    Parameters:
    - name (str): Full name to be split.

    Returns:
    tuple: A tuple containing the first name and last name.
    """
    # Split the name string to 2 parts.
    nameParts = name.split(" ", 1)
    firstName = nameParts[0]
    lastName = nameParts[1] if len(nameParts) > 1 else ""

    return firstName, lastName


def linkedInLogin():
    """
    Perform LinkedIn login using a WebDriver.

    This function initializes a WebDriver, navigates to the LinkedIn login page, enters the provided
    username and password, and logs into the LinkedIn account.

    Returns:
    selenium.webdriver.chrome.webdriver.WebDriver: The WebDriver object after successful login.
    """
    driver = None
    try:
        driver = webdriver.Chrome()
        linkedinUrl = "https://www.linkedin.com/"
        driver.get(linkedinUrl)
        time.sleep(2)
    except NoSuchDriverException as e:
        print(f"No Such Driver\n{e}")
    if driver:
        try:
            # Finding the username and the password elements on the driver screen
            username = driver.find_element(By.XPATH, "//input[@name='session_key']")
            password = driver.find_element(By.XPATH, "//input[@name='session_password']")

            # Wait until the finalSignIn button is visible, located and clickable (up to 20 seconds)
            finalSignInBtn = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.CLASS_NAME, "sign-in-form__submit-btn--full-width")) and
                EC.presence_of_element_located((By.CLASS_NAME, "sign-in-form__submit-btn--full-width")) and
                EC.element_to_be_clickable((By.CLASS_NAME, "sign-in-form__submit-btn--full-width"))
            )

            # Getting the username and password from the environment variables of the system (we don't want to explicitly place it in the script).
            username.send_keys(os.environ.get("LinkedInUserName"))
            password.send_keys(os.environ.get("LinkedInPassword"))

            # Wait until the username and password is visible in the right input fields and are correct (up to 20 seconds).
            WebDriverWait(driver, 20).until(
                EC.text_to_be_present_in_element_value((By.XPATH, "//input[@name='session_key']"),
                                                       os.environ.get("LinkedInUserName")) and
                EC.text_to_be_present_in_element_value((By.XPATH, "//input[@name='session_password']"),
                                                       os.environ.get("LinkedInPassword"))
            )
            time.sleep(2)
            # Press the Sign-In button and log in.
            finalSignInBtn.click()
            time.sleep(4)
        except TimeoutException as e:
            print(f"Exception in function 'linkedInLogin'\n{e}")

        # returns the driver object.
        return driver


def openLink(driver, linkedInUrl):
    """
    Open a LinkedIn profile link using a WebDriver.

    This function navigates the provided WebDriver to the specified LinkedIn profile URL.

    Parameters:
    - driver (selenium.webdriver.chrome.webdriver.WebDriver): The WebDriver object.
    - linkedInUrl (str): The LinkedIn profile URL to be opened.

    Returns:
    None
    """
    try:
        driver.get(linkedInUrl)
        time.sleep(3)
    except WebDriverException as e:
        print(f"Exception in function 'openLink'\n{e}")


def findButtonFromType(driver, btn_type, click_=True):
    """
    Open the LinkedIn user message box using a WebDriver.

    This function waits for the 'Message' button to become visible, then clicks on it to open the message box
    for sending a message to a LinkedIn user.

    Parameters:
    - driver (selenium.webdriver.chrome.webdriver.WebDriver): The WebDriver object.

    Returns:
    None
    :param click_:
    """
    try:
        time.sleep(3)
        # Finding all the buttons on the person page
        all_buttons = driver.find_elements(By.TAG_NAME, "button")
        # Select only the 'Message' button
        messageBtn = [btn for btn in all_buttons if btn.text == btn_type]
        time.sleep(2)
        # print("messageBtn is ", len(messageBtn))
        # Preform the click on the 'Message' button.
        if click_:
            messageBtn[0].click()
            return True
    except ElementNotVisibleException as e:
        print(f"Exception in function 'openLinkedInUserMessageBox'\n{e}")
        return False
    except IndexError as e:
        print(f"Exception in function 'openLinkedInUserMessageBox'\n{e}\n"
              f"Index out of bound because button elements haven't been loaded in time.")
        return False

    except Exception as e:
        print(f"Exception in function 'openLinkedInUserMessageBox'\n{e}")
        return False


def clickMessageArea(driver):
    """
    Click on the message area of the LinkedIn user message box using a WebDriver.

    This function waits for the message area to become visible, then clicks on it to set focus for entering a message.

    Parameters:
    - driver (selenium.webdriver.chrome.webdriver.WebDriver): The WebDriver object.

    Returns:
    None
    """
    try:
        time.sleep(2)
        messageAreaMainDiv = driver.find_element(By.XPATH,
                                                 "//div[starts-with(@class, 'msg-form__msg-content-container')]")
        messageAreaMainDiv.click()
        time.sleep(2)
    except NoSuchElementException as e:
        print(f"Exception in function 'clickMessageArea'\n{e}")


def findMessageParagraphAndEnterMessageTemplet(driver, message):
    """
    Find a paragraph in the LinkedIn user message box and enter the provided message using a WebDriver.

    This function waits for paragraphs to be present on the page, then enters the provided message into the input field.

    Parameters:
    - driver (selenium.webdriver.chrome.webdriver.WebDriver): The WebDriver object.
    - message (str): The message to be entered into the LinkedIn user message box.

    Returns:
    None
    """
    try:
        # Wait for the paragraphs to be present on the page
        paragraphs = WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "p"))
        )

        # Enter the message to the input field (assuming paragraphs[-5] is the correct element)
        paragraphs[-5].send_keys(message)

    except TimeoutException as e:
        print(f"Exception in function 'clickMessageArea'\nTimeout waiting for paragraphs to be present: {e}")

    except NoSuchElementException as e:
        print(f"Exception in function 'clickMessageArea'\nParagraphs not found on the page: {e}")


def sendMessage(driver):
    """
    Send a message in the LinkedIn user message box using a WebDriver.

    This function finds the 'Send' button, clicks it to send the message, and then closes the chat window.

    Parameters:
    - driver (selenium.webdriver.chrome.webdriver.WebDriver): The WebDriver object.

    Returns:
    None
    """
    try:
        time.sleep(2)
        # Find the 'Send' button.
        sendMessageBtn = driver.find_element(By.CLASS_NAME, "msg-form__send-button")

        # Send the message.
        sendMessageBtn.click()
        time.sleep(3)
        # Find all buttons.
        all_buttons = driver.find_elements(By.TAG_NAME, "button")
        # Find the close conversation button.
        closeConversationBtn = [btn for btn in all_buttons if
                                "Close your conversation with" in btn.text or "Close your" in btn.text]
        time.sleep(3)
        # Close the chat window.
        closeConversationBtn[0].click()
    except ElementNotInteractableException as e:
        print(f"Exception in function 'sendMessage'\n{e}")

    except TimeoutException as e:
        print(f"Exception in function 'sendMessage'\nTimeout waiting for an element: {e}")


def getCurrentTime():
    """
    Get the current time in the format HH:MM:SS.

    Returns:
    str: A string representing the current time.
    """
    t = time.localtime()
    current_time = time.strftime("%H:%M:%S", t)
    return current_time


def getDateAndTime():
    """
    Get the current date and time in the format dd/mm/YYYY HH:MM:SS.

    Returns:
    str: A string representing the current date and time.
    """
    # datetime object containing current date and time
    now = datetime.now()

    # dd/mm/YY H:M:S
    dtString = now.strftime("%d/%m/%Y %H:%M:%S")
    return dtString


def timeToNextMessagingRound(seconds, exitKey):
    """
    Wait for a specified number of seconds, displaying a countdown, and check for an exit signal.

    Args:
    seconds (int): The duration to wait in seconds.
    exitKey (str): The keyboard shortcut to check for an exit signal.

    Returns:
    bool: True if the exit signal is detected during the countdown, False otherwise.
    """
    for i in range(seconds, 0, -1):
        hours, remainder = divmod(i, 3600)
        minutes, seconds = divmod(remainder, 60)

        sys.stdout.write(f"\r{hours:02d}:{minutes:02d}:{seconds:02d}")
        sys.stdout.flush()

        time.sleep(1)
        if checkExitProgram(exitKey):
            return True
    return False


def checkExitProgram(exitKey):
    """
    Check if a specified keyboard shortcut is pressed, indicating a request to exit the program.

    Args:
    exitKey (str): The keyboard shortcut to check for an exit signal.

    Returns:
    bool: True if the exit signal is detected, False otherwise.
    """
    if keyboard.is_pressed(exitKey):
        print(f"\nYou pressed {exitKey}. Exiting...")
        return True
    return False


def getContent(html, container_type, name):
    """
    Extract content from HTML based on container type and keyword found in text, id, or class.

    :param html: The HTML string to parse.
    :param container_type: The tag name to search for (e.g., 'div', 'section').
    :param name: The keyword to search for in the text, id, or class attributes.
    :param exact_match: Whether to match the keyword exactly (default: False).
    :param case_insensitive: Whether to ignore case (default: True).
    :return: The cleaned HTML of the matching container(s) or "None" if no match is found.
    """
    from bs4 import BeautifulSoup

    if not html or not container_type:
        return "Invalid input"

    soup = BeautifulSoup(html, 'html.parser')
    containers = soup.find_all(container_type)
    matches = []

    for container in containers:
        filtered_html = clean_html(filter_html(container.prettify()))
        if name in filtered_html.lower():
            if len(matches) > 2:
                break
            matches.append(filtered_html)

    return '\n'.join(matches) if matches else "None"


import re


def filter_html(html):
    """
    Remove HTML markup, inline JavaScript/CSS, comments, and unnecessary whitespace
    from the given HTML string.

    :param html: The HTML string to be cleaned.
    :type html: str
    :return: A cleaned string without HTML tags or unnecessary content.
    :rtype: str
    """

    # Remove inline JavaScript and CSS (e.g., <script>...</script>, <style>...</style>)
    cleaned = re.sub(r"(?is)<(script|style).*?>.*?(</\1>)", "", html.strip())
    # Remove HTML comments (e.g., <!-- comment -->)
    cleaned = re.sub(r"(?s)<!--.*?-->", "", cleaned)
    # Remove all remaining HTML tags (e.g., <div>, <a>)
    cleaned = re.sub(r"(?s)<.*?>", " ", cleaned)
    # Replace common HTML entities with spaces
    cleaned = re.sub(r"&nbsp;", " ", cleaned)
    # Remove repeated spaces
    cleaned = re.sub(r"\s+", " ", cleaned)
    cleaned = cleaned.replace("\n", " ")
    # Remove unwanted phrases (e.g., Follow, Connect, Message)
    cleaned = re.sub(r'\b(Follow|Connect|Message|Save to PDF|More)\b', '', cleaned, flags=re.IGNORECASE)
    # Trim and return the cleaned text
    return cleaned.strip()
